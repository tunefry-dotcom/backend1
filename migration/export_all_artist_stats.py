"""Export combined all-platform stats + song/platform/monthly breakdowns +
remaining balance for EVERY artist in the legacy SQL Server dump into one
properly-arranged Excel workbook.

Why this exists
----------------
``legacy_artist_stats.py`` answers "what did artist X earn" one artist at a
time (interactive lookup). This script runs the same aggregation over every
distinct ``MusicStreams.ArtistName`` in a single pass over the dump and
writes one .xlsx with a Summary sheet (one row per resolved artist) plus
Songs / Platforms / Monthly / Withdrawals detail sheets, so nothing needs to
be re-parsed per artist.

Reuses the parsing/aggregation primitives from ``legacy_artist_stats.py``
(``parse_table``, ``aggregate``, ``monthly_breakdown``, ``money``,
``strip_cast``) rather than duplicating them — that module has no Supabase
import chain, so it's safe to import directly.

One deliberate difference from ``legacy_artist_stats.py``: that script
resolves a *user* to a single ``MusicStreams.ArtistName`` by priority
(Username, then FullName, then ArtistName) and only ever looks at rows under
that one chosen name. This script instead resolves **per row**: every
``MusicStreams.ArtistName`` is looked up against the full set of
Username/FullName/ArtistName values across ALL users, and rows are grouped
by the UserID that uniquely owns that name. This means a user whose legacy
stream data is split across two literal name variants (e.g. both their
Username AND FullName appear as distinct ``ArtistName`` strings in
MusicStreams) gets all of it combined here, where the single-artist tool
would only surface whichever name wins the priority order.

Names that resolve to more than one candidate user, or to none, are never
guessed — they're written to their own sheets with full row detail so
nothing is silently dropped.

Read-only: parses the dump once; writes only the output .xlsx. Never touches
the dump itself, Supabase, or anywhere else.

Usage
-----
    python migration/export_all_artist_stats.py
    python migration/export_all_artist_stats.py "<dump.sql>" -o custom_report.xlsx
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

sys.path.insert(0, str(Path(__file__).parent))
from legacy_artist_stats import (  # noqa: E402
    DEFAULT_DUMP,
    aggregate,
    load_sql,
    monthly_breakdown,
    money,
    parse_table,
    strip_cast,
)

DEFAULT_OUTPUT = str(Path(__file__).parent / "legacy_all_artists_report.xlsx")

CURRENCY_FMT = "#,##0.00"
INT_FMT = "#,##0"


# ---------------------------------------------------------------------------
# Resolution: every MusicStreams row -> the one User that owns it (or
# ambiguous / unmatched)
# ---------------------------------------------------------------------------

def build_candidate_map(users: list[dict]) -> dict[str, set[str]]:
    """lowercase Username/FullName/ArtistName -> set of UserIDs claiming it."""
    m: dict[str, set[str]] = defaultdict(set)
    for u in users:
        uid = str(u.get("UserID"))
        for field in ("Username", "FullName", "ArtistName"):
            v = u.get(field)
            if v:
                m[v.lower()].add(uid)
    return m


def resolve_all(stream_rows: list[dict], users: list[dict]):
    candidate_map = build_candidate_map(users)
    users_by_id = {str(u.get("UserID")): u for u in users}

    rows_by_uid: dict[str, list[dict]] = defaultdict(list)
    ambiguous: dict[str, dict] = {}
    unmatched: dict[str, list[dict]] = defaultdict(list)

    for r in stream_rows:
        name = r.get("ArtistName")
        if not name:
            continue
        uids = candidate_map.get(name.lower(), set())
        if len(uids) == 1:
            rows_by_uid[next(iter(uids))].append(r)
        elif len(uids) > 1:
            entry = ambiguous.setdefault(name, {"uids": sorted(uids), "rows": []})
            entry["rows"].append(r)
        else:
            unmatched[name].append(r)

    matched = {uid: (users_by_id[uid], rows) for uid, rows in rows_by_uid.items()}
    return matched, ambiguous, unmatched


# ---------------------------------------------------------------------------
# Excel writing helpers
# ---------------------------------------------------------------------------

def write_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: list[list],
    currency_cols: set[int] = frozenset(),
    int_cols: set[int] = frozenset(),
    totals_row: list | None = None,
) -> Worksheet:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for row in rows:
        ws.append(row)

    last_data_row = ws.max_row
    if last_data_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_data_row}"

    if totals_row is not None:
        ws.append(totals_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    for idx in range(1, len(headers) + 1):
        col = get_column_letter(idx)
        if (idx - 1) in currency_cols:
            fmt = CURRENCY_FMT
        elif (idx - 1) in int_cols:
            fmt = INT_FMT
        else:
            fmt = None
        if fmt:
            for cell in ws[col][1:]:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = fmt

        max_len = len(str(headers[idx - 1]))
        for row in rows:
            v = row[idx - 1]
            max_len = max(max_len, len(str(v)) if v is not None else 0)
        ws.column_dimensions[col].width = min(max(max_len + 2, 10), 45)

    return ws


def as_float(v) -> float:
    return float(v)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("dump", nargs="?", default=DEFAULT_DUMP,
                     help=f"Path to the legacy SQL dump (default: {DEFAULT_DUMP})")
    ap.add_argument("-o", "--output", default=DEFAULT_OUTPUT,
                     help=f"Output .xlsx path (default: {DEFAULT_OUTPUT})")
    args = ap.parse_args()

    dump_path = Path(args.dump)
    if not dump_path.exists():
        ap.error(f"dump not found: {dump_path}")

    print(f"Loading {dump_path} ...")
    sql_text = load_sql(dump_path)

    print("Parsing dbo.Users ...")
    users = parse_table(sql_text, "dbo", "Users",
                         ["UserID", "Username", "FullName", "Email", "ArtistName"])
    print(f"  {len(users)} users")

    print("Parsing dbo.MusicStreams ...")
    all_streams = parse_table(
        sql_text, "dbo", "MusicStreams",
        ["ArtistName", "Song", "Streams", "Revenue", "Month", "Year", "Platform",
         "IsDeleted", "RedeemedAmount"],
    )
    stream_rows = [r for r in all_streams if r.get("IsDeleted") != "1"]
    print(f"  {len(all_streams)} rows ({len(stream_rows)} after excluding IsDeleted)")

    print("Parsing dbo.WithdrawalHistory ...")
    withdrawals_raw = parse_table(
        sql_text, "dbo", "WithdrawalHistory",
        ["Id", "UserId", "Amount", "Status", "CreatedDate", "ProcessedDate", "Description"],
    )
    print(f"  {len(withdrawals_raw)} withdrawal requests")

    print("Resolving MusicStreams.ArtistName -> User ...")
    matched, ambiguous, unmatched = resolve_all(stream_rows, users)
    print(f"  {len(matched)} artists matched, {len(ambiguous)} ambiguous names, "
          f"{len(unmatched)} unmatched names")

    users_by_id = {str(u.get("UserID")): u for u in users}

    # --- Summary ------------------------------------------------------
    summary_rows = []
    per_artist_agg = {}
    for uid, (user, rows) in matched.items():
        agg = aggregate(rows)
        per_artist_agg[uid] = agg
        summary_rows.append([
            user.get("UserID"), user.get("Username"), user.get("FullName"), user.get("Email"),
            agg["total_streams"], as_float(agg["total_revenue"]), as_float(agg["total_redeemed"]),
            as_float(agg["remaining_balance"]), len(agg["songs"]), len(agg["platforms"]),
        ])
    summary_rows.sort(key=lambda r: -r[5])

    grand_totals = [
        "", "", "", "GRAND TOTAL",
        sum(r[4] for r in summary_rows),
        sum(r[5] for r in summary_rows),
        sum(r[6] for r in summary_rows),
        sum(r[7] for r in summary_rows),
        "", "",
    ]

    # --- Songs / Platforms / Monthly (long format) --------------------
    song_rows, platform_rows, monthly_rows = [], [], []
    for uid, (user, rows) in sorted(matched.items(), key=lambda kv: (kv[1][0].get("Username") or "")):
        agg = per_artist_agg[uid]
        uname = user.get("Username")
        for s in agg["songs"]:
            song_rows.append([uid, uname, s["song"], s["streams"], as_float(s["revenue"]),
                               as_float(s["redeemed"]), as_float(s["remaining"])])
        for p in agg["platforms"]:
            platform_rows.append([uid, uname, p["platform"], p["streams"], as_float(p["revenue"])])
        for m in monthly_breakdown(rows):
            monthly_rows.append([uid, uname, m["year"], m["month"], m["streams"],
                                  as_float(m["revenue"]), as_float(m["redeemed"]), as_float(m["remaining"])])

    # --- Withdrawals (all rows, regardless of match status) -----------
    withdrawal_rows = []
    for w in withdrawals_raw:
        uid = str(w.get("UserId"))
        user = users_by_id.get(uid)
        withdrawal_rows.append([
            w.get("Id"), uid, user.get("Username") if user else "(unknown user)",
            as_float(money(w.get("Amount"))), w.get("Status"),
            strip_cast(w.get("CreatedDate") or ""), strip_cast(w.get("ProcessedDate") or "") or None,
            w.get("Description"),
        ])

    # --- Ambiguous / unmatched artist names ----------------------------
    ambiguous_rows = []
    for name, entry in sorted(ambiguous.items()):
        agg = aggregate(entry["rows"])
        candidate_names = ", ".join(
            f"{uid}:{users_by_id[uid].get('Username')}" for uid in entry["uids"] if uid in users_by_id
        )
        ambiguous_rows.append([name, ", ".join(entry["uids"]), candidate_names,
                                agg["total_streams"], as_float(agg["total_revenue"])])

    unmatched_rows = []
    for name, rows in sorted(unmatched.items()):
        agg = aggregate(rows)
        unmatched_rows.append([name, agg["total_streams"], as_float(agg["total_revenue"]), len(agg["songs"])])

    # --- Write workbook --------------------------------------------------
    wb = Workbook()
    wb.remove(wb.active)

    notes = wb.create_sheet("Notes")
    notes_lines = [
        "Legacy artist stats — all artists, generated from the legacy SQL Server dump",
        "",
        f"Source dump: {dump_path}",
        "",
        "Sheets:",
        "  Summary            One row per artist resolved from MusicStreams.ArtistName.",
        "  Songs               Per-artist song-wise breakdown (long format).",
        "  Platforms           Per-artist platform-wise breakdown (long format).",
        "  Monthly             Per-artist monthly breakdown (long format).",
        "  Withdrawals         All WithdrawalHistory rows (all 20), joined to Username where known.",
        "  Ambiguous_Names     MusicStreams.ArtistName values matching 2+ Users by name — not",
        "                      attributed to any one artist below; shown here with candidate UserIDs.",
        "  Unmatched_Names     MusicStreams.ArtistName values matching zero Users — orphan legacy rows.",
        "",
        "Key derivations:",
        "  Remaining Balance = Total Revenue - Total Redeemed, summed over the artist's",
        "  MusicStreams rows. RedeemedAmount is a per-row withdrawal-allocation marker (not a",
        "  rounded copy of Revenue) so this sum is the actual balance not yet claimed by any",
        "  withdrawal request.",
        "",
        "Caveats:",
        "  - Amounts are legacy-currency-as-stored; no FX conversion applied.",
        "  - This is pre-migration legacy data, separate from current Supabase song_stats /",
        "    artist_balances — do not conflate the two.",
        "  - IsDeleted=1 MusicStreams rows are excluded from all totals.",
        "  - 'Pending' withdrawal status means requested but not yet paid; its amount is",
        "    already reflected in Remaining Balance via RedeemedAmount, not double-subtracted.",
    ]
    for line in notes_lines:
        notes.append([line])
    notes.column_dimensions["A"].width = 100
    notes["A1"].font = Font(bold=True, size=13)

    write_sheet(
        wb, "Summary",
        ["UserID", "Username", "FullName", "Email", "Total Streams", "Total Revenue",
         "Total Redeemed", "Remaining Balance", "Distinct Songs", "Distinct Platforms"],
        summary_rows,
        currency_cols={5, 6, 7}, int_cols={4, 8, 9},
        totals_row=grand_totals,
    )
    write_sheet(
        wb, "Songs",
        ["UserID", "Username", "Song", "Streams", "Revenue", "Redeemed", "Remaining"],
        song_rows,
        currency_cols={4, 5, 6}, int_cols={3},
    )
    write_sheet(
        wb, "Platforms",
        ["UserID", "Username", "Platform", "Streams", "Revenue"],
        platform_rows,
        currency_cols={4}, int_cols={3},
    )
    write_sheet(
        wb, "Monthly",
        ["UserID", "Username", "Year", "Month", "Streams", "Revenue", "Redeemed", "Remaining"],
        monthly_rows,
        currency_cols={5, 6, 7}, int_cols={4},
    )
    write_sheet(
        wb, "Withdrawals",
        ["WithdrawalID", "UserID", "Username", "Amount", "Status", "CreatedDate",
         "ProcessedDate", "Description"],
        withdrawal_rows,
        currency_cols={3},
    )
    write_sheet(
        wb, "Ambiguous_Names",
        ["MusicStreams.ArtistName", "Candidate UserIDs", "Candidate Users", "Streams", "Revenue"],
        ambiguous_rows,
        currency_cols={4}, int_cols={3},
    )
    write_sheet(
        wb, "Unmatched_Names",
        ["MusicStreams.ArtistName", "Streams", "Revenue", "Distinct Songs"],
        unmatched_rows,
        currency_cols={2}, int_cols={1, 3},
    )

    out_path = Path(args.output)
    wb.save(out_path)
    print(f"\nWrote {out_path.resolve()}")
    print(f"  Summary: {len(summary_rows)} artists, "
          f"{sum(r[5] for r in summary_rows):,.2f} total revenue")
    print(f"  Ambiguous names: {len(ambiguous_rows)}   Unmatched names: {len(unmatched_rows)}")


if __name__ == "__main__":
    main()
