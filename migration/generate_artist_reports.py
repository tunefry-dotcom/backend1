"""Generate per-individual-artist Excel reports from Royalty_Report.xlsx,
enriched with account identity from a legacy reference file and SQL dump.

Attribution rules for multi-artist rows:
  - Exactly 1 credited artist has a Tunefry account → only that artist gets the royalty.
  - 2+ credited artists all have Tunefry accounts → CONFLICT: all receive the royalty
    and the row is logged to the Conflicts sheet for manual review.
  - 0 credited artists have Tunefry accounts → all name-only artists receive the royalty.
  - Single-artist rows → always attributed normally.

Identity resolution order:
  1. Reference file Summary sheet (legacy_all_artists_report.xlsx) → "Reference File"
  2. SQL dump dbo.Users (table with data.sql)                      → "Legacy SQL Dump"
  3. Not found anywhere                                             → "Name Only"

total_balance = remaining_balance = total_royalty  (no withdrawal deductions).

Usage:
    python migration/generate_artist_reports.py
    python migration/generate_artist_reports.py "path/Royalty_Report.xlsx" \\
        --reference "path/legacy_all_artists_report.xlsx" \\
        --output "path/Artist_Reports.xlsx"
"""
from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

_HERE = Path(__file__).parent
sys.path.insert(0, str(_HERE))

try:
    from legacy_artist_stats import load_sql, parse_table
except ImportError as exc:
    sys.exit(f"Cannot import from legacy_artist_stats.py: {exc}")

try:
    from platform_map import normalize_platform
except ImportError as exc:
    sys.exit(f"Cannot import from platform_map.py: {exc}")

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required.  Run: pip install openpyxl")

# ── defaults ──────────────────────────────────────────────────────────────────

DEFAULT_DUMP      = r"C:\Users\ViditVaibhav\Downloads\table with data.sql"
DEFAULT_REPORT    = r"C:\Users\ViditVaibhav\Desktop\Royalty_Report.xlsx"
DEFAULT_REFERENCE = r"C:\Users\ViditVaibhav\Desktop\tunefry reports\old reports\legacy_all_artists_report.xlsx"
DEFAULT_OUTPUT    = r"C:\Users\ViditVaibhav\Desktop\Artist_Reports.xlsx"

_MULTI_ARTIST_RE = re.compile(
    r'\s{2,}|,\s*|\s+(?:featuring|feat\.?|ft\.?|x|&|and)\s+',
    re.IGNORECASE,
)

_MONTH_NAMES = {
    "01": "January",   "02": "February",  "03": "March",
    "04": "April",     "05": "May",       "06": "June",
    "07": "July",      "08": "August",    "09": "September",
    "10": "October",   "11": "November",  "12": "December",
}

# ── Excel style helpers ───────────────────────────────────────────────────────

_DARK_BLUE   = "1F4E79"
_MID_BLUE    = "2E74B5"
_LIGHT_BLUE  = "BDD7EE"
_WHITE       = "FFFFFF"
_ORANGE      = "C55A11"
_LIGHT_ORANGE = "FCE4D6"
_INR_FMT     = '#,##0.00'


def _write_header_row(ws, row: int, cols: list[str], color: str = _DARK_BLUE) -> None:
    fill = PatternFill("solid", fgColor=color)
    font = Font(bold=True, color=_WHITE)
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _section_title(ws, row: int, title: str, ncols: int = 6) -> None:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, color=_WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=_MID_BLUE)
    if ncols > 1:
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=ncols,
        )


def _kv(ws, row: int, label: str, value, source_color: bool = False) -> None:
    lc = ws.cell(row=row, column=1, value=label)
    lc.font = Font(bold=True)
    lc.fill = PatternFill("solid", fgColor=_LIGHT_BLUE)
    vc = ws.cell(row=row, column=2, value=value)
    if source_color and isinstance(value, str):
        if value == "Reference File":
            vc.font = Font(color="375623", bold=True)
        elif value == "Legacy SQL Dump":
            vc.font = Font(color="833C00", bold=True)
        else:
            vc.font = Font(color="595959")


def _autofit(ws) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                widths[cell.column] = max(
                    widths.get(cell.column, 8),
                    min(len(str(cell.value)) + 4, 52),
                )
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _f(val: Decimal) -> float:
    return float(val)

# ── identity loading ──────────────────────────────────────────────────────────


def load_reference_identities(path: Path) -> dict[str, dict]:
    """Read Summary sheet from legacy_all_artists_report.xlsx → name → identity dict."""
    if not path.exists():
        print(f"  WARNING: Reference file not found at {path} — skipping.")
        return {}

    print(f"Loading reference file: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Summary" not in wb.sheetnames:
        print("  WARNING: No 'Summary' sheet in reference file — skipping.")
        wb.close()
        return {}

    ws = wb["Summary"]
    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, ())
    headers = [str(h or "").strip() for h in raw_headers]

    lookup: dict[str, dict] = {}
    count = 0
    for raw in rows_iter:
        if not any(v is not None for v in raw):
            continue
        rd = dict(zip(headers, raw))
        uid      = rd.get("UserID")
        username = str(rd.get("Username") or "").strip()
        fullname = str(rd.get("FullName") or "").strip()
        email    = str(rd.get("Email") or "").strip()
        if not username and not uid:
            continue
        identity = {
            "UserID":   uid,
            "Username": username,
            "FullName": fullname,
            "Email":    email,
            "source":   "Reference File",
        }
        for field_val in (username, fullname):
            if field_val:
                lookup.setdefault(field_val.lower(), identity)
        count += 1

    wb.close()
    print(f"  {count} reference accounts loaded ({len(lookup)} name keys).")
    return lookup


def load_sql_identities(dump_path: Path) -> dict[str, dict]:
    """Parse dbo.Users from SQL dump → name → identity dict."""
    print(f"Loading SQL dump: {dump_path}")
    print("  (UTF-16 file — may take 30-60 s ...)")
    sql_text = load_sql(dump_path)
    print("  Parsing dbo.Users ...")
    users = parse_table(
        sql_text, "dbo", "Users",
        ["UserID", "Username", "FullName", "ArtistName", "Email"],
    )
    print(f"  {len(users)} user rows found.")

    lookup: dict[str, dict] = {}
    for u in users:
        identity = {
            "UserID":     u.get("UserID"),
            "Username":   str(u.get("Username") or "").strip(),
            "FullName":   str(u.get("FullName") or "").strip(),
            "ArtistName": str(u.get("ArtistName") or "").strip(),
            "Email":      str(u.get("Email") or "").strip(),
            "source":     "Legacy SQL Dump",
        }
        for field in ("ArtistName", "Username", "FullName"):
            raw = str(u.get(field) or "").strip()
            if raw:
                lookup.setdefault(raw.lower(), identity)
    return lookup

# ── royalty report reading ────────────────────────────────────────────────────


def _col(headers: list[str], *candidates: str) -> str | None:
    for c in candidates:
        if c in headers:
            return c
    return None


def _as_dec(v) -> Decimal:
    if v is None:
        return Decimal("0")
    try:
        return Decimal(str(v).replace(",", "").strip() or "0")
    except InvalidOperation:
        return Decimal("0")


def _as_int(v) -> int:
    try:
        return int(float(str(v or 0).replace(",", "") or 0))
    except (ValueError, TypeError):
        return 0


def read_report(path: Path, sheet_override: str | None) -> list[dict]:
    print(f"\nReading: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    if sheet_override:
        if sheet_override not in wb.sheetnames:
            sys.exit(f"Sheet '{sheet_override}' not found in workbook.")
        ws = wb[sheet_override]
    elif "Combined_All" in wb.sheetnames:
        ws = wb["Combined_All"]
        print("  Using sheet: Combined_All")
    else:
        ws = wb[wb.sheetnames[0]]
        print(f"  Using first sheet: {wb.sheetnames[0]}")

    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, ())
    headers = [str(h or "").strip().lower() for h in raw_headers]

    royalty_c   = _col(headers, "royalty_inr", "royalty_inr_amount", "royalty")
    artist_c    = _col(headers, "artist", "artist_name", "track_artist")
    title_c     = _col(headers, "track_title", "song_name", "asset_title", "title", "song")
    platform_c  = _col(headers, "source_platform", "platform", "service", "dsp")
    qty_c       = _col(headers, "quantity", "total", "streams", "qty")
    period_c    = _col(headers, "source_period", "period")
    isrc_c      = _col(headers, "isrc")
    sub_label_c = _col(headers, "sub_label", "sublabel")
    adj_c       = _col(headers, "adjustment_type", "adjustment")

    print(
        f"  Columns — royalty={royalty_c!r}  artist={artist_c!r}  "
        f"title={title_c!r}  platform={platform_c!r}  qty={qty_c!r}  "
        f"period={period_c!r}  sub_label={sub_label_c!r}"
    )

    def g(rd: dict, col: str | None) -> str:
        return str(rd.get(col, "") or "").strip() if col else ""

    rows: list[dict] = []
    skipped = 0
    for raw in rows_iter:
        if not any(v is not None for v in raw):
            continue
        rd = dict(zip(headers, raw))
        if adj_c and str(rd.get(adj_c, "") or "").strip() == "Dispute Resolution":
            skipped += 1
            continue
        rows.append({
            "artist":    g(rd, artist_c),
            "title":     g(rd, title_c) or "(Unknown)",
            "platform":  g(rd, platform_c),
            "qty":       _as_int(rd.get(qty_c)),
            "royalty":   _as_dec(rd.get(royalty_c)),
            "period":    g(rd, period_c),
            "isrc":      g(rd, isrc_c),
            "sub_label": g(rd, sub_label_c),
        })

    wb.close()
    print(f"  {len(rows)} data rows loaded.  {skipped} Dispute Resolution rows skipped.")
    return rows

# ── identity resolution ───────────────────────────────────────────────────────


def _exact(name: str, lookup: dict) -> dict | None:
    return lookup.get(name.lower())


def _substring(name: str, lookup: dict) -> dict | None:
    key = name.lower().strip()
    if len(key) < 4:
        return None
    for stored, identity in lookup.items():
        if key in stored or stored in key:
            return identity
    return None


def resolve_identity(name: str, ref: dict, sql: dict) -> dict:
    identity = _exact(name, ref) or _exact(name, sql) or _substring(name, sql)
    if identity:
        return identity
    return {
        "UserID": "", "Username": "", "FullName": "",
        "ArtistName": "", "Email": "", "source": "Name Only",
    }


def is_registered(identity: dict) -> bool:
    """True if this artist has a Tunefry account (Reference File or SQL Dump)."""
    return identity.get("source") in ("Reference File", "Legacy SQL Dump")


def collect_unique_artists(rows: list[dict]) -> set[str]:
    """First-pass scan: collect every unique individual artist name across all rows."""
    names: set[str] = set()
    for row in rows:
        artist_raw = row["artist"] or "(Unknown)"
        for p in _MULTI_ARTIST_RE.split(artist_raw):
            p = p.strip()
            if p:
                names.add(p)
    return names


def resolve_all_identities(
    names: set[str],
    ref: dict[str, dict],
    sql: dict[str, dict],
) -> dict[str, dict]:
    """Resolve identity for every unique artist name upfront (needed before aggregation)."""
    return {name: resolve_identity(name, ref, sql) for name in names}

# ── aggregation ───────────────────────────────────────────────────────────────


def _new_group(name: str, identity: dict) -> dict:
    return {
        "name":          name,
        "identity":      identity,
        "streams":       0,
        "royalty":       Decimal("0"),
        "has_conflicts": False,
        "songs":         defaultdict(lambda: {
            "streams": 0,
            "royalty": Decimal("0"),
            "artists": set(),
            "distributors": set(),
        }),
        "platforms":     defaultdict(lambda: {"streams": 0, "royalty": Decimal("0")}),
        "months":        defaultdict(lambda: {"streams": 0, "royalty": Decimal("0")}),
    }


def _split_artists(artist_raw: str) -> list[str]:
    parts = _MULTI_ARTIST_RE.split(artist_raw)
    seen_lower: set[str] = set()
    result = []
    for p in parts:
        p = p.strip()
        if p and p.lower() not in seen_lower:
            result.append(p)
            seen_lower.add(p.lower())
    return result or [artist_raw]


def aggregate(
    rows: list[dict],
    name_to_identity: dict[str, dict],
) -> tuple[dict[str, dict], list[dict]]:
    """
    Aggregate rows by individual artist with Tunefry-aware attribution:
      single-artist row     → attributed normally regardless of registration
      multi-artist, 1 reg   → only the registered Tunefry artist gets the royalty
      multi-artist, 2+ reg  → CONFLICT: all registered artists receive royalty + row flagged
      multi-artist, 0 reg   → all name-only artists receive royalty (no account dispute)

    Returns (groups dict, list of conflict dicts).
    """
    groups: dict[str, dict] = {}
    # conflict key: (artist_raw_lower, title_lower)
    conflicts_map: dict[tuple, dict] = {}

    multi_rows = 0
    single_attribution = 0
    conflict_rows = 0
    name_only_multi = 0

    for row in rows:
        artist_raw = row["artist"] or "(Unknown)"
        individuals = _split_artists(artist_raw)

        qty     = row["qty"]
        royalty = row["royalty"]
        title   = row["title"]
        sub     = row["sub_label"] or "(no label)"
        _, pg   = normalize_platform(row["platform"])
        p_str   = row["period"]
        mk      = p_str if re.match(r"\d{4}-\d{2}", p_str) else (p_str or "Unknown")

        is_conflict = False

        if len(individuals) == 1:
            # Single artist — always attributed
            targets = individuals
        else:
            multi_rows += 1
            registered = [
                a for a in individuals
                if is_registered(name_to_identity.get(a, {}))
            ]

            if len(registered) == 1:
                # Exactly one Tunefry account — only they get the royalty
                targets = registered
                single_attribution += 1
            elif len(registered) >= 2:
                # Multiple Tunefry accounts on one row — conflict
                targets = registered
                is_conflict = True
                conflict_rows += 1
                ck = (artist_raw.lower(), title.lower())
                if ck not in conflicts_map:
                    conflicts_map[ck] = {
                        "artist_raw":   artist_raw,
                        "title":        title,
                        "conflicting":  registered,
                        "sub_label":    sub,
                        "royalty":      Decimal("0"),
                        "periods":      set(),
                    }
                conflicts_map[ck]["royalty"] += royalty
                conflicts_map[ck]["periods"].add(mk)
            else:
                # No Tunefry accounts — attribute to all name-only
                targets = individuals
                name_only_multi += 1

        for artist_name in targets:
            if artist_name not in groups:
                idt = name_to_identity.get(artist_name, {
                    "UserID": "", "Username": "", "FullName": "",
                    "ArtistName": "", "Email": "", "source": "Name Only",
                })
                groups[artist_name] = _new_group(artist_name, idt)
            g = groups[artist_name]

            if is_conflict:
                g["has_conflicts"] = True

            g["streams"]             += qty
            g["royalty"]              += royalty
            g["songs"][title]["streams"]      += qty
            g["songs"][title]["royalty"]       += royalty
            g["songs"][title]["artists"].add(artist_raw)
            g["songs"][title]["distributors"].add(sub)
            g["platforms"][pg]["streams"]     += qty
            g["platforms"][pg]["royalty"]      += royalty
            g["months"][mk]["streams"]         += qty
            g["months"][mk]["royalty"]          += royalty

    print(f"\n  Attribution breakdown (multi-artist rows only):")
    print(f"    Multi-artist rows total:          {multi_rows}")
    print(f"    Attributed to 1 registered only:  {single_attribution}")
    print(f"    Conflict (2+ registered):         {conflict_rows}")
    print(f"    All name-only (0 registered):     {name_only_multi}")

    return groups, list(conflicts_map.values())

# ── workbook builder ──────────────────────────────────────────────────────────


def _safe_sheet_name(name: str, existing: list[str]) -> str:
    base = re.sub(r'[\\/*?\[\]:]', '_', (name or "Sheet").strip())[:31]
    candidate, n = base, 2
    while candidate in existing:
        suffix = f"_{n}"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    return candidate


def _write_artist_sheet(wb: Workbook, gd: dict) -> None:
    identity = gd["identity"] or {}
    src = identity.get("source", "Name Only")

    def iv(field: str) -> str:
        return str(identity.get(field) or "")

    sheet_name = _safe_sheet_name(gd["name"], [ws.title for ws in wb.worksheets])
    ws = wb.create_sheet(sheet_name)
    r = 1

    # Conflict warning banner
    if gd["has_conflicts"]:
        cell = ws.cell(row=r, column=1,
            value="CONFLICT: One or more collaborative tracks credit multiple Tunefry accounts. "
                  "See the Conflicts sheet for details.")
        cell.font = Font(bold=True, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor=_ORANGE)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

    # ── Identity ──────────────────────────────────────────────────────────────
    _section_title(ws, r, "ARTIST IDENTITY", ncols=4); r += 1
    _kv(ws, r, "Artist Name",    gd["name"]); r += 1
    _kv(ws, r, "Identity Source", src, source_color=True); r += 1
    _kv(ws, r, "UserID",          iv("UserID") or "—"); r += 1
    _kv(ws, r, "Username",        iv("Username") or "—"); r += 1
    _kv(ws, r, "Full Name",       iv("FullName") or "—"); r += 1
    _kv(ws, r, "Email",           iv("Email") or "—"); r += 1
    r += 1

    # ── Overall stats ─────────────────────────────────────────────────────────
    balance = gd["royalty"]
    _section_title(ws, r, "OVERALL STATS", ncols=4); r += 1
    for lbl, val, fmt in [
        ("Total Streams",           gd["streams"],    None),
        ("Total Royalty (INR)",     _f(balance),      _INR_FMT),
        ("Total Balance (INR)",     _f(balance),      _INR_FMT),
        ("Remaining Balance (INR)", _f(balance),      _INR_FMT),
    ]:
        lc = ws.cell(row=r, column=1, value=lbl)
        lc.font = Font(bold=True)
        lc.fill = PatternFill("solid", fgColor="BDD7EE")
        vc = ws.cell(row=r, column=2, value=val)
        if fmt:
            vc.number_format = fmt
        r += 1
    r += 1

    # ── Song breakdown ────────────────────────────────────────────────────────
    _section_title(ws, r, "SONG BREAKDOWN", ncols=6); r += 1
    _write_header_row(ws, r, [
        "Song", "Full Artist Credit", "Distributor", "Streams", "Royalty (INR)", "Balance (INR)",
    ]); r += 1
    for song, sv in sorted(gd["songs"].items(), key=lambda kv: kv[1]["royalty"], reverse=True):
        artists_label = " / ".join(sorted(sv["artists"]))
        dist_label    = " / ".join(sorted(sv["distributors"]))
        ws.cell(r, 1, song)
        ws.cell(r, 2, artists_label)
        ws.cell(r, 3, dist_label)
        ws.cell(r, 4, sv["streams"])
        c5 = ws.cell(r, 5, _f(sv["royalty"]));  c5.number_format = _INR_FMT
        c6 = ws.cell(r, 6, _f(sv["royalty"]));  c6.number_format = _INR_FMT
        r += 1
    r += 1

    # ── Platform breakdown ────────────────────────────────────────────────────
    _section_title(ws, r, "PLATFORM BREAKDOWN", ncols=3); r += 1
    _write_header_row(ws, r, ["Platform", "Streams", "Royalty (INR)"]); r += 1
    for pg, pv in sorted(gd["platforms"].items(), key=lambda kv: kv[1]["royalty"], reverse=True):
        ws.cell(r, 1, pg)
        ws.cell(r, 2, pv["streams"])
        c = ws.cell(r, 3, _f(pv["royalty"]));  c.number_format = _INR_FMT
        r += 1
    r += 1

    # ── Monthly breakdown ─────────────────────────────────────────────────────
    _section_title(ws, r, "MONTHLY BREAKDOWN (Feb-Apr 2026)", ncols=4); r += 1
    _write_header_row(ws, r, ["Year", "Month", "Streams", "Royalty (INR)"]); r += 1
    for mk in sorted(gd["months"].keys()):
        mv = gd["months"][mk]
        if re.match(r"\d{4}-\d{2}$", mk):
            yr, mo = mk.split("-")
            month_label = _MONTH_NAMES.get(mo, mo)
        else:
            yr, month_label = "", mk
        ws.cell(r, 1, yr)
        ws.cell(r, 2, month_label)
        ws.cell(r, 3, mv["streams"])
        c = ws.cell(r, 4, _f(mv["royalty"]));  c.number_format = _INR_FMT
        r += 1

    _autofit(ws)


def _write_conflicts_sheet(wb: Workbook, conflicts: list[dict]) -> None:
    """Add a Conflicts sheet listing all rows where 2+ Tunefry accounts share credit."""
    ws = wb.create_sheet("Conflicts")
    ws.sheet_properties.tabColor = "C55A11"  # orange tab — stands out among blue artist tabs
    r = 1

    note = ws.cell(row=r, column=1,
        value=(
            f"CONFLICT REVIEW REQUIRED — {len(conflicts)} track(s) credit multiple Tunefry "
            "accounts. Each conflicted track's royalty has been attributed to ALL credited "
            "registered accounts. Manually decide the correct attribution and adjust."
        ))
    note.font = Font(bold=True, color=_WHITE)
    note.fill = PatternFill("solid", fgColor=_ORANGE)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2

    _write_header_row(ws, r, [
        "Track Title", "Original Artist String",
        "Conflicting Tunefry Accounts",
        "Distributor", "Total Royalty (INR)", "Periods",
    ]); r += 1

    for c in sorted(conflicts, key=lambda x: x["royalty"], reverse=True):
        ws.cell(r, 1, c["title"])
        ws.cell(r, 2, c["artist_raw"])
        ws.cell(r, 3, ", ".join(c["conflicting"]))
        ws.cell(r, 4, c["sub_label"])
        vc = ws.cell(r, 5, _f(c["royalty"]));  vc.number_format = _INR_FMT
        ws.cell(r, 6, ", ".join(sorted(c["periods"])))
        # Highlight the row
        fill = PatternFill("solid", fgColor="FCE4D6")
        for col in range(1, 7):
            ws.cell(r, col).fill = fill
        r += 1

    _autofit(ws)


def build_workbook(
    groups: dict[str, dict],
    conflicts: list[dict],
    report_total: Decimal,
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    sorted_groups = sorted(groups.values(), key=lambda g: g["royalty"], reverse=True)

    # Count identity sources
    ref_count  = sum(1 for g in sorted_groups if g["identity"].get("source") == "Reference File")
    sql_count  = sum(1 for g in sorted_groups if g["identity"].get("source") == "Legacy SQL Dump")
    name_only  = sum(1 for g in sorted_groups if g["identity"].get("source") == "Name Only")
    n_conflict = sum(1 for g in sorted_groups if g["has_conflicts"])

    # ── Conflicts sheet FIRST so it's always tab #1 ───────────────────────────
    _write_conflicts_sheet(wb, conflicts)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_properties.tabColor = "1F4E79"  # dark blue tab

    note_cell = ws_sum.cell(row=1, column=1,
        value=(
            f"Royalty report total (INR): {float(report_total):,.2f}  |  "
            f"Artists: {len(sorted_groups)}  |  "
            f"Reference File: {ref_count}  |  "
            f"SQL Dump: {sql_count}  |  "
            f"Name Only: {name_only}  |  "
            f"Conflict flags: {n_conflict}  |  "
            "NOTE: Multi-artist rows attributed only to registered Tunefry accounts where possible."
        ))
    note_cell.font = Font(italic=True, color="595959")
    ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

    _write_header_row(ws_sum, 2, [
        "#", "Artist Name", "UserID", "Username", "Full Name", "Email",
        "Identity Source", "Conflict?", "Total Streams",
        "Total Royalty (INR)", "Remaining Balance (INR)",
    ])
    ws_sum.freeze_panes = "A3"

    for idx, gd in enumerate(sorted_groups, 1):
        idt  = gd["identity"] or {}
        bal  = _f(gd["royalty"])
        src  = idt.get("source", "Name Only")
        ws_sum.append([
            idx,
            gd["name"],
            str(idt.get("UserID") or ""),
            str(idt.get("Username") or ""),
            str(idt.get("FullName") or ""),
            str(idt.get("Email") or ""),
            src,
            "YES" if gd["has_conflicts"] else "",
            gd["streams"],
            bal,
            bal,
        ])
        row_num = idx + 2
        for col_i in (10, 11):
            ws_sum.cell(row=row_num, column=col_i).number_format = _INR_FMT
        # Colour the source cell
        src_cell = ws_sum.cell(row=row_num, column=7)
        if src == "Reference File":
            src_cell.font = Font(color="375623", bold=True)
        elif src == "Legacy SQL Dump":
            src_cell.font = Font(color="833C00")
        else:
            src_cell.font = Font(color="595959")
        # Colour conflict cell
        if gd["has_conflicts"]:
            c_cell = ws_sum.cell(row=row_num, column=8)
            c_cell.font = Font(color=_ORANGE, bold=True)

    _autofit(ws_sum)

    # ── Per-artist sheets ─────────────────────────────────────────────────────
    print(f"  Writing {len(sorted_groups)} artist sheets ...")
    for i, gd in enumerate(sorted_groups, 1):
        _write_artist_sheet(wb, gd)
        if i % 200 == 0:
            print(f"    ... {i}/{len(sorted_groups)} sheets written")

    return wb

# ── CLI ───────────────────────────────────────────────────────────────────────


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Generate per-individual-artist Excel reports from a royalty report, "
            "enriched with identity from a reference file and legacy SQL dump."
        ),
    )
    ap.add_argument("report", nargs="?", default=DEFAULT_REPORT)
    ap.add_argument("--dump",      default=DEFAULT_DUMP)
    ap.add_argument("--reference", default=DEFAULT_REFERENCE)
    ap.add_argument("--output",    default=DEFAULT_OUTPUT)
    ap.add_argument("--sheet",     default=None)
    args = ap.parse_args()

    report_path    = Path(args.report)
    dump_path      = Path(args.dump)
    reference_path = Path(args.reference)
    output_path    = Path(args.output)

    if not report_path.exists():
        sys.exit(f"Royalty report not found: {report_path}")
    if not dump_path.exists():
        sys.exit(f"SQL dump not found: {dump_path}")

    # Load identity lookups
    ref_identities = load_reference_identities(reference_path)
    sql_identities = load_sql_identities(dump_path)
    print(f"  SQL name lookup: {len(sql_identities)} entries")

    # Read royalty report
    rows = read_report(report_path, args.sheet)
    report_total = sum(r["royalty"] for r in rows)
    print(f"  Report total (INR): {float(report_total):,.2f}")

    # Pre-resolve ALL identities before aggregation
    print("\nPre-resolving identities for all unique artist names ...")
    all_names = collect_unique_artists(rows)
    print(f"  Unique artist names found: {len(all_names)}")
    name_to_identity = resolve_all_identities(all_names, ref_identities, sql_identities)

    ref_pre  = sum(1 for v in name_to_identity.values() if v["source"] == "Reference File")
    sql_pre  = sum(1 for v in name_to_identity.values() if v["source"] == "Legacy SQL Dump")
    none_pre = sum(1 for v in name_to_identity.values() if v["source"] == "Name Only")
    print(f"    Reference File: {ref_pre}  |  SQL Dump: {sql_pre}  |  Name Only: {none_pre}")

    # Aggregate
    print("\nAggregating (Tunefry-aware attribution) ...")
    groups, conflicts = aggregate(rows, name_to_identity)

    grand_sum = sum(g["royalty"] for g in groups.values())
    print(f"\nResults:")
    print(f"  Unique artists with royalty:  {len(groups)}")
    print(f"  Report total (INR):           {float(report_total):,.2f}")
    print(f"  Sum across artist sheets:     {float(grand_sum):,.2f}")
    print(f"  Conflict tracks flagged:      {len(conflicts)}")

    # Build and save
    print("\nBuilding workbook ...")
    wb = build_workbook(groups, conflicts, report_total)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    sheet_count = len(wb.worksheets)
    print(f"\nSaved: {output_path}")
    print(f"  {sheet_count} sheets  (Summary + Conflicts + {sheet_count - 2} artist sheets)")


if __name__ == "__main__":
    main()
