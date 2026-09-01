"""Monthly ingest of the DSP-consolidated royalty report -> earnings tables.

Reads a Combined_All sheet from a Tunefry ``Royalty_Detail_Reports_*.xlsx``
workbook and updates:

  * ``public.song_stats``       — per (email x song x platform x month) rows
                                  (period-replace strategy: delete existing
                                  rows for the covered periods per matched
                                  user, then insert fresh — safe re-runs)
  * ``public.artist_balances``  — recomputed from the full song_stats table
                                  plus ``migration/withdrawn_baseline.json``
                                  and the current Supabase paid/pending
                                  withdrawal_requests state
  * ``public.submissions``      — sets status='approved' on any submission
                                  whose ISRC appears in the file (never
                                  demotes an existing approved row)

Money-critical invariants
-------------------------
* ``migration/withdrawn_baseline.json`` MUST exist. It is the immutable
  snapshot of each artist's legacy withdrawn (tunefry adjustments +
  legacy WithdrawalHistory) that this script cannot re-derive. Regenerate
  ONLY via ``compute_withdrawn_baseline.py``.
* ``migration/fx_rates.json`` MUST contain a rate for EVERY ``source_period``
  present in the Excel file. We refuse to silently zero-convert.
* ``song_stats.period_month`` is written as the FULL English month name
  (``"February"``, not ``"02"`` or ``"Feb"``) — required by the frontend
  chart month order.
* Currency: values in ``royalty`` (Excel, USD) are multiplied by the FX
  rate for that period and stored as INR at Decimal(20,10) precision.
* Plan royalty % is NOT applied — matches ``ingest_streams.py`` behavior.
* Rows with ``adjustment_type == "Dispute Resolution"`` are excluded from
  both streams and revenue totals.

Attribution
-----------
Excel has no user email. We match on artist name:
  1. Legacy artist -> email map (from public.profiles.artist_name +
     auth.users.raw_user_meta_data.artist_name). 1-to-1 mappings only —
     names shared by multiple users are dropped.
  2. Fallback: sub_label -> email via the same map (label owners often
     have their label name as sub_label in the report).

Unmatched artists are written to ``unmatched_<timestamp>.csv`` alongside the
workbook path. No rows are written for them until you fix the ``artist_name``
in ``profiles`` and re-run.

Usage
-----
    # Dry run (always start here):
    python migration/ingest_royalty_report.py "<path\\to\\report.xlsx>" \\
        --fx-rates migration/fx_rates.json \\
        --dry-run

    # Live run (writes to Supabase):
    python migration/ingest_royalty_report.py "<path\\to\\report.xlsx>" \\
        --fx-rates migration/fx_rates.json
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

sys.path.insert(0, ".")
from dotenv import load_dotenv; load_dotenv()

import openpyxl

from app.core.supabase_client import get_service_client
from migration.platform_map import normalize_platform


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
BASELINE_PATH = Path("migration/withdrawn_baseline.json")
DEFAULT_FX_PATH = Path("migration/fx_rates.json")
LEGACY_MAP_PATH = Path("migration/legacy_map.json")
UPSERT_BATCH = 500
DELETE_CHUNK = 200

MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------
def norm(s: str | None) -> str:
    """Loose key for artist / title matching (lowercase, whitespace-collapsed)."""
    return re.sub(r"\s+", " ", (str(s) if s is not None else "").strip().lower())


def to_decimal(v) -> Decimal:
    try:
        s = str(v).strip() if v is not None else "0"
        if s == "" or s.lower() in ("none", "nan", "null"):
            return Decimal("0")
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def to_int(v) -> int:
    try:
        if v is None or v == "":
            return 0
        return int(float(v))
    except (ValueError, TypeError):
        return 0


def month_name_from_period(period: str | None) -> tuple[str, int] | None:
    """('2026-02', ...) -> ('February', 2026). None on garbage input."""
    if not period:
        return None
    m = re.match(r"^(\d{4})-(\d{2})$", str(period).strip())
    if not m:
        return None
    year = int(m.group(1))
    idx = int(m.group(2))
    if not 1 <= idx <= 12:
        return None
    return (MONTH_NAMES[idx - 1], year)


# ---------------------------------------------------------------------------
# Config loaders
# ---------------------------------------------------------------------------
def load_fx_rates(path: Path) -> dict[str, Decimal]:
    """Load {"2026-02": 90.0, ...} -> {"2026-02": Decimal("90.0")}."""
    if not path.exists():
        print(f"ERROR: FX rates file not found: {path}", file=sys.stderr)
        sys.exit(2)
    data = json.loads(path.read_text(encoding="utf-8"))
    rates: dict[str, Decimal] = {}
    for k, v in data.items():
        if k.startswith("_"):
            continue
        try:
            rates[k] = Decimal(str(v))
        except InvalidOperation:
            print(f"ERROR: FX rates: bad value for {k!r}: {v!r}", file=sys.stderr)
            sys.exit(2)
    return rates


def load_baseline() -> dict[str, Decimal]:
    if not BASELINE_PATH.exists():
        print(f"ERROR: {BASELINE_PATH} missing. Run "
              f"migration/compute_withdrawn_baseline.py first.", file=sys.stderr)
        sys.exit(2)
    raw = json.loads(BASELINE_PATH.read_text(encoding="utf-8"))
    return {k.lower(): Decimal(str(v)) for k, v in raw.items()}


def load_legacy_map() -> dict:
    """Load migration/legacy_map.json if present, else return empty maps.

    Produced by migration/build_legacy_map.py from the legacy SQL dump.
    Contains authoritative artist_name/full_name/username/ISRC/UPC/(artist,title)
    -> email mappings, filtered to Supabase-existing accounts.
    """
    empty = {"artist_name": {}, "full_name": {}, "username": {},
             "isrc": {}, "upc": {}, "artist_title": {}}
    if not LEGACY_MAP_PATH.exists():
        print(f"WARN: {LEGACY_MAP_PATH} missing — attribution will be less "
              f"accurate. Run migration/build_legacy_map.py to build it.")
        return empty
    raw = json.loads(LEGACY_MAP_PATH.read_text(encoding="utf-8"))
    return {
        "artist_name": {k: v.lower() for k, v in (raw.get("artist_name") or {}).items()},
        "full_name":   {k: v.lower() for k, v in (raw.get("full_name") or {}).items()},
        "username":    {k: v.lower() for k, v in (raw.get("username") or {}).items()},
        "isrc":        {k.upper(): v.lower() for k, v in (raw.get("isrc") or {}).items()},
        "upc":         {k.upper(): v.lower() for k, v in (raw.get("upc") or {}).items()},
        "artist_title": {k: v.lower() for k, v in (raw.get("artist_title") or {}).items()},
    }


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------
EXPECTED_COLUMNS = [
    "row_id", "source_platform", "source_period", "source_month", "source_file",
    "service", "period_start", "period_end", "country",
    "isrc", "upc", "ean", "grid", "release_id", "custom_id",
    "artist", "track_title", "album_release",
    "label", "main_label", "sub_label", "label_identification",
    "product_service", "subscription_type", "content_format", "company",
    "distribution_id", "original_l1_name", "original_l2_name",
    "quantity",
    # USD workbook columns (Royalty_Detail_Reports_*.xlsx):
    "income", "admin_exp", "royalty",
    # INR workbook columns (Royalty_Reports_*_INR.xlsx — pre-converted):
    "source_currency", "income_source", "admin_exp_source",
    "royalty_source", "fx_to_inr", "income_inr", "admin_exp_inr", "royalty_inr",
    "native_sheet", "native_row",
    # optional / present on some sheets:
    "adjustment_type",
]


def read_workbook(path: Path, sheet_name: str) -> list[dict]:
    print(f"Opening workbook: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: sheet {sheet_name!r} not in workbook. Available: "
              f"{wb.sheetnames}", file=sys.stderr)
        sys.exit(2)
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    headers = [(h or "").strip().lower() if isinstance(h, str) else h for h in header]
    out: list[dict] = []
    for row in rows_iter:
        if row is None:
            continue
        rec = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        # Skip fully-blank rows (openpyxl sometimes yields these at tail).
        if not any(v not in (None, "") for v in rec.values()):
            continue
        out.append(rec)
    wb.close()
    print(f"  {len(out)} data rows in sheet {sheet_name!r}")
    return out


# ---------------------------------------------------------------------------
# Apple Music INR CSV reader
# ---------------------------------------------------------------------------
def read_apple_csv(path: Path, period_override: str | None = None) -> tuple[list[dict], str]:
    """Read an Apple Music INR CSV (revenue already in INR, no FX needed).

    Columns used: isrc, item_artist, song_name, total (streams),
    royality (net INR revenue), Sub_Label.
    Period auto-detected from filename like ``_03_2026_`` (MM_YYYY) or
    supplied via ``period_override`` as a ``YYYY-MM`` string.

    Returns ``(rows, period_str)``.  Each row is a normalised dict ready
    for the secondary aggregation loop.  Revenue is already INR — the main
    loop must NOT apply an FX rate to these rows.
    """
    period = period_override
    if not period:
        m = re.search(r"_(\d{2})_(\d{4})", path.stem)
        if m:
            period = f"{m.group(2)}-{m.group(1)}"
    if not period:
        print(f"ERROR: Cannot auto-detect period from CSV filename {path.name!r}. "
              f"Pass --csv-period YYYY-MM.", file=sys.stderr)
        sys.exit(2)
    if not re.match(r"^\d{4}-\d{2}$", period):
        print(f"ERROR: --csv-period must be YYYY-MM, got {period!r}", file=sys.stderr)
        sys.exit(2)
    if month_name_from_period(period) is None:
        print(f"ERROR: --csv-period {period!r} is not a valid YYYY-MM period.",
              file=sys.stderr)
        sys.exit(2)

    with open(path, newline="", encoding="utf-8-sig") as f:
        raw_rows = list(csv.DictReader(f))
    out: list[dict] = []
    for r in raw_rows:
        out.append({
            "source_platform": "Apple Music",
            "source_period":   period,
            "isrc":            (r.get("isrc") or "").strip().upper(),
            "artist":          (r.get("item_artist") or "").strip(),
            "track_title":     (r.get("song_name") or "").strip(),
            "sub_label":       (r.get("Sub_Label") or r.get("sub_label") or "").strip(),
            "quantity":        r.get("total") or "0",
            "_royalty_inr":    to_decimal(r.get("royality") or "0"),
        })
    print(f"  {len(out)} rows from Apple Music CSV (period={period})")
    return out, period


# ---------------------------------------------------------------------------
# Attribution
# ---------------------------------------------------------------------------
def build_artist_email_map(svc) -> tuple[dict[str, str], int]:
    """Build normalized-artist-name -> email map from profiles + auth users.

    Returns (map, dropped_ambiguous_count). Names mapping to more than one
    email are dropped to avoid misattributing money.
    """
    name_emails: dict[str, set[str]] = defaultdict(set)

    # profiles.artist_name
    offset = 0
    while True:
        res = (svc.table("profiles")
               .select("user_id, artist_name, full_name")
               .range(offset, offset + 999).execute())
        page = res.data or []
        for r in page:
            uid = r.get("user_id")
            if not uid:
                continue
            # We look up email in auth.users below by iterating listUsers;
            # here we just collect (name -> user_id) pairs.
            for field in ("artist_name", "full_name"):
                n = norm(r.get(field))
                if n:
                    name_emails[n].add(uid)
        if len(page) < 1000:
            break
        offset += 1000

    # Resolve user_ids -> emails via auth.admin.list_users (paginated).
    uid_to_email: dict[str, str] = {}
    page_num = 1
    while True:
        page = svc.auth.admin.list_users(page=page_num, per_page=1000)
        # supabase-py returns a list here (Python SDK 2.x).
        users = page if isinstance(page, list) else getattr(page, "users", []) or []
        if not users:
            break
        for u in users:
            uid = getattr(u, "id", None) or (u.get("id") if isinstance(u, dict) else None)
            email = getattr(u, "email", None) or (u.get("email") if isinstance(u, dict) else None)
            if uid and email:
                uid_to_email[str(uid)] = email.lower()
            # Also index by user_metadata.artist_name / full_name
            meta = getattr(u, "user_metadata", None)
            if meta is None and isinstance(u, dict):
                meta = u.get("user_metadata")
            meta = meta or {}
            for field in ("artist_name", "full_name"):
                n = norm(meta.get(field) if isinstance(meta, dict) else None)
                if n and email:
                    name_emails[n].add(str(uid) if uid else email.lower())
        if len(users) < 1000:
            break
        page_num += 1

    # Now dedupe: a name is usable iff it maps to exactly ONE email.
    # (Multiple user_ids -> same email is fine; multiple emails is not.)
    final_map: dict[str, str] = {}
    dropped = 0
    for name, ids in name_emails.items():
        emails = {uid_to_email.get(x, x) if x in uid_to_email else x for x in ids}
        # Some entries were emails already (auth-only path). Filter to real emails.
        emails = {e.lower() for e in emails if isinstance(e, str) and "@" in e}
        if len(emails) == 1:
            final_map[name] = next(iter(emails))
        elif len(emails) > 1:
            dropped += 1
    return final_map, dropped


def build_isrc_submission_map(svc) -> dict[tuple[str, str], str]:
    """(email, isrc) -> submission_id.

    Scans all submissions with any ISRC in ``data``. Uses:
      - data->>isrc          (legacy migrated rows)
      - data->>isrc_code     (transfer-song / transfer-album top-level)
      - data->'songs'[N]->>isrc  (transfer-album per-track)
    """
    out: dict[tuple[str, str], str] = {}
    offset = 0
    while True:
        res = (svc.table("submissions")
               .select("id, user_email, data")
               .range(offset, offset + 999).execute())
        page = res.data or []
        for r in page:
            sub_id = r.get("id")
            email = (r.get("user_email") or "").lower()
            data = r.get("data") or {}
            isrcs: set[str] = set()
            for k in ("isrc", "isrc_code"):
                v = data.get(k)
                if isinstance(v, str) and v.strip():
                    isrcs.add(v.strip().upper())
            for track in (data.get("songs") or []):
                if isinstance(track, dict):
                    v = track.get("isrc") or track.get("isrc_code")
                    if isinstance(v, str) and v.strip():
                        isrcs.add(v.strip().upper())
            if not email or not sub_id:
                continue
            for isrc in isrcs:
                out.setdefault((email, isrc), sub_id)
        if len(page) < 1000:
            break
        offset += 1000
    return out


def build_name_submission_map(svc) -> dict[tuple[str, str], str]:
    """(email, norm_song_title) -> submission_id. Fallback when ISRC misses."""
    out: dict[tuple[str, str], str] = {}
    offset = 0
    while True:
        res = (svc.table("submissions")
               .select("id, user_email, data")
               .range(offset, offset + 999).execute())
        page = res.data or []
        for r in page:
            sub_id = r.get("id")
            email = (r.get("user_email") or "").lower()
            data = r.get("data") or {}
            if not email or not sub_id:
                continue
            titles: set[str] = set()
            for k in ("song_title", "songTitle", "song_name", "title"):
                v = data.get(k)
                if isinstance(v, str) and v.strip():
                    titles.add(norm(v))
            for track in (data.get("songs") or []):
                if isinstance(track, dict):
                    v = track.get("title") or track.get("song_title")
                    if isinstance(v, str) and v.strip():
                        titles.add(norm(v))
            for t in titles:
                out.setdefault((email, t), sub_id)
        if len(page) < 1000:
            break
        offset += 1000
    return out


def load_reserved_withdrawals(svc) -> tuple[dict[str, Decimal], dict[str, Decimal]]:
    """(paid_by_email, pending_by_email) — current Supabase state."""
    paid: dict[str, Decimal] = defaultdict(Decimal)
    pending: dict[str, Decimal] = defaultdict(Decimal)
    offset = 0
    while True:
        try:
            res = (svc.table("withdrawal_requests")
                   .select("user_email, amount, status")
                   .range(offset, offset + 999).execute())
        except Exception:
            break
        page = res.data or []
        for r in page:
            email = (r.get("user_email") or "").lower()
            amt = to_decimal(r.get("amount"))
            if r.get("status") == "paid":
                paid[email] += amt
            elif r.get("status") == "pending":
                pending[email] += amt
        if len(page) < 1000:
            break
        offset += 1000
    return paid, pending


def list_existing_balance_emails(svc) -> set[str]:
    """All emails currently in artist_balances — so we recompute them all."""
    out: set[str] = set()
    offset = 0
    while True:
        res = (svc.table("artist_balances").select("user_email")
               .range(offset, offset + 999).execute())
        page = res.data or []
        for r in page:
            e = (r.get("user_email") or "").lower()
            if e:
                out.add(e)
        if len(page) < 1000:
            break
        offset += 1000
    return out


# ---------------------------------------------------------------------------
# Unmatched Excel report
# ---------------------------------------------------------------------------
def _write_unmatched_excel(
    path: Path,
    unmatched_artists: dict[str, dict],
    is_inr_workbook: bool,
) -> None:
    """Write a 3-sheet Excel of unmatched artists alongside the CSV.

    Sheets:
      Artists   — one row per artist, colour-coded by INR revenue tier
      Songs     — one row per (artist, track), sorted by INR revenue desc
      Platforms — aggregated platform breakdown across all unmatched artists
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    fill_red    = PatternFill("solid", fgColor="FFB3B3")   # >= INR 10 000
    fill_yellow = PatternFill("solid", fgColor="FFF2CC")   # >= INR 1 000
    fill_header = PatternFill("solid", fgColor="4472C4")
    font_header = Font(bold=True, color="FFFFFF")
    align_ctr   = Alignment(horizontal="center")

    sort_key = (
        (lambda a: -(float(a.get("total_royalty_inr") or 0)))
        if is_inr_workbook
        else (lambda a: -(float(a.get("total_royalty_usd") or 0)))
    )
    sorted_artists = sorted(unmatched_artists.values(), key=sort_key)

    def _set_header(ws, headers: list[str]) -> None:
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = font_header
            c.fill = fill_header
            c.alignment = align_ctr
        ws.freeze_panes = "A2"

    def _autofit(ws) -> None:
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
            ws.column_dimensions[
                get_column_letter(col_cells[0].column)
            ].width = min(max_len + 4, 60)

    # ---- Sheet 1: Artists ----
    ws1 = wb.active
    ws1.title = "Artists"
    h1 = ["#", "Artist", "Sub Label", "INR Revenue", "Streams",
          "Songs", "Platforms", "Periods"]
    _set_header(ws1, h1)

    for rank, a in enumerate(sorted_artists, 1):
        inr  = float(a.get("total_royalty_inr") or 0)
        plat = ", ".join(sorted(a.get("platform_data", {}).keys()))
        per  = ", ".join(sorted(a.get("periods", set())))
        n_songs = len(a.get("tracks", {}))
        row_data = [rank, a["artist"], a.get("sub_label") or "",
                    round(inr, 2), a.get("total_streams") or 0,
                    n_songs, plat, per]
        ri = rank + 1
        for col, val in enumerate(row_data, 1):
            ws1.cell(row=ri, column=col, value=val)
        fill = fill_red if inr >= 10000 else (fill_yellow if inr >= 1000 else None)
        if fill:
            for col in range(1, len(h1) + 1):
                ws1.cell(row=ri, column=col).fill = fill

    _autofit(ws1)

    # ---- Sheet 2: Songs ----
    ws2 = wb.create_sheet("Songs")
    h2 = ["Artist", "Track Title", "ISRC", "INR Revenue", "Streams"]
    _set_header(ws2, h2)

    all_tracks: list[tuple[str, dict]] = []
    for a in sorted_artists:
        for tk in a.get("tracks", {}).values():
            all_tracks.append((a["artist"], tk))
    all_tracks.sort(key=lambda t: -(float(t[1].get("inr") or 0)))

    for ri, (artist, tk) in enumerate(all_tracks, 2):
        ws2.cell(row=ri, column=1, value=artist)
        ws2.cell(row=ri, column=2, value=tk.get("title") or "")
        ws2.cell(row=ri, column=3, value=tk.get("isrc") or "")
        ws2.cell(row=ri, column=4, value=round(float(tk.get("inr") or 0), 2))
        ws2.cell(row=ri, column=5, value=tk.get("streams") or 0)

    _autofit(ws2)

    # ---- Sheet 3: Platforms ----
    ws3 = wb.create_sheet("Platforms")
    h3 = ["Platform", "INR Revenue", "Streams", "Artists", "Tracks"]
    _set_header(ws3, h3)

    plat_agg: dict[str, dict] = {}
    for a in sorted_artists:
        for plat, pd in a.get("platform_data", {}).items():
            if plat not in plat_agg:
                plat_agg[plat] = {"inr": Decimal("0"), "streams": 0,
                                   "artists": set(), "tracks": 0}
            plat_agg[plat]["inr"]     += pd["inr"]
            plat_agg[plat]["streams"] += pd["streams"]
            plat_agg[plat]["artists"].add(a["artist"])
            plat_agg[plat]["tracks"]  += len(a.get("tracks", {}))

    for ri, (plat, pd) in enumerate(
        sorted(plat_agg.items(), key=lambda kv: -(float(kv[1]["inr"]))), 2
    ):
        ws3.cell(row=ri, column=1, value=plat)
        ws3.cell(row=ri, column=2, value=round(float(pd["inr"]), 2))
        ws3.cell(row=ri, column=3, value=pd["streams"])
        ws3.cell(row=ri, column=4, value=len(pd["artists"]))
        ws3.cell(row=ri, column=5, value=pd["tracks"])

    _autofit(ws3)

    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("xlsx", help="Path to the Royalty_Detail_Reports_*.xlsx workbook")
    ap.add_argument("--fx-rates", default=str(DEFAULT_FX_PATH),
                    help="Path to fx_rates.json (default: migration/fx_rates.json)")
    ap.add_argument("--sheet", default="Combined_All",
                    help="Sheet to read (default: Combined_All)")
    # SAFETY: dry-run is the DEFAULT. You must pass --live to actually write.
    ap.add_argument("--live", action="store_true",
                    help="Perform live writes to Supabase. Without this, dry-run only.")
    ap.add_argument("--dry-run", action="store_true",
                    help="(default) Compute + print everything without touching the DB")
    ap.add_argument("--csv",
                    help="Apple Music INR CSV path (revenue already in INR; no FX applied). "
                         "Platform is fixed to 'Apple Music'. For the overlapping period, "
                         "the CSV data fully replaces the xlsx Apple Music rows.")
    ap.add_argument("--csv-period",
                    help="Period for --csv as YYYY-MM (auto-detected from filename if omitted).")
    ap.add_argument("--expected-gross", dest="expected_gross",
                    help="Expected total gross INR from the report, e.g. '147075.74'. "
                         "Shown in reconciliation summary for operator review.")
    ap.add_argument("--expected-net", dest="expected_net",
                    help="Expected total net royalty INR, e.g. '126361.52'. "
                         "Ingest ABORTS (dry-run: warns) if total to write exceeds this.")
    args = ap.parse_args()
    # Effective dry_run: True unless --live was passed.
    dry_run: bool = not args.live

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERROR: workbook not found: {xlsx_path}", file=sys.stderr)
        sys.exit(2)

    fx_rates = load_fx_rates(Path(args.fx_rates))
    baseline = load_baseline()
    legacy = load_legacy_map()
    print(f"FX rates loaded: {sorted(fx_rates.keys())}")
    print(f"Baseline loaded: {len(baseline)} users, "
          f"sum(legacy_withdrawn)={sum(baseline.values())}")
    print(f"Legacy map: {len(legacy['artist_name'])} artist names, "
          f"{len(legacy['full_name'])} full names, "
          f"{len(legacy['username'])} usernames, "
          f"{len(legacy['isrc'])} ISRCs, {len(legacy['upc'])} UPCs, "
          f"{len(legacy['artist_title'])} (artist||title)")

    rows = read_workbook(xlsx_path, args.sheet)
    if not rows:
        print("No data rows — nothing to ingest.")
        return

    # ---- Load Apple Music CSV (supplementary INR data) ------------------
    csv_rows: list[dict] = []
    csv_period: str | None = None
    total_csv_inr_all = Decimal("0")
    if args.csv:
        csv_file = Path(args.csv)
        if not csv_file.exists():
            print(f"ERROR: CSV not found: {csv_file}", file=sys.stderr)
            sys.exit(2)
        print(f"\nLoading Apple Music CSV: {csv_file}")
        csv_rows, csv_period = read_apple_csv(csv_file, args.csv_period)
        if not csv_rows:
            print("  Apple Music CSV is empty — ignoring.")
            csv_period = None
        else:
            total_csv_inr_all = sum(r["_royalty_inr"] for r in csv_rows)
            print(f"  Apple Music CSV total net INR: ₹{total_csv_inr_all:.2f}")

    # Detect whether this is a pre-converted INR workbook (royalty_inr column
    # present) or the original USD workbook (royalty column, needs FX).
    is_inr_workbook = any(
        r.get("royalty_inr") is not None
        for r in rows[:min(100, len(rows))]
    )
    if is_inr_workbook:
        print("[INFO] INR workbook detected (royalty_inr column, fx_to_inr=1)"
              " — no FX conversion will be applied.")
    else:
        print("[INFO] USD workbook detected (royalty column) — FX rates will be applied.")

    # Sanity: every source_period in the file must have an FX rate (USD mode only).
    file_periods = {str(r.get("source_period") or "").strip()
                    for r in rows if r.get("source_period")}
    if not is_inr_workbook:
        missing = sorted(p for p in file_periods if p not in fx_rates)
        if missing:
            print(f"ERROR: fx_rates.json is missing entries for: {missing}",
                  file=sys.stderr)
            print("       Add them and re-run. We never silently zero-convert.",
                  file=sys.stderr)
            sys.exit(2)
    print(f"File covers periods: {sorted(file_periods)}")

    # ---- Attribution maps -----------------------------------------------
    svc = get_service_client()
    print("Building artist -> email map…")
    artist_to_email, dropped = build_artist_email_map(svc)
    supa_map_size = len(artist_to_email)
    # Merge legacy SQL-dump-derived name maps. Legacy takes precedence when
    # both sources have the same normalized name — SQL dump is the
    # authoritative source for who registered which artist name.
    for src in (legacy["artist_name"], legacy["full_name"], legacy["username"]):
        for name, email in src.items():
            artist_to_email[name] = email
    print(f"  {supa_map_size} from Supabase + {len(artist_to_email) - supa_map_size} "
          f"new from legacy = {len(artist_to_email)} total names -> email "
          f"({dropped} Supabase-side ambiguous names dropped)")

    print("Building ISRC -> submission_id map…")
    isrc_to_sub = build_isrc_submission_map(svc)
    print(f"  {len(isrc_to_sub)} (email, ISRC) pairs indexed")

    print("Building (song title) -> submission_id fallback map…")
    title_to_sub = build_name_submission_map(svc)
    print(f"  {len(title_to_sub)} (email, title) pairs indexed")

    # Reverse lookups: ISRC-only / title-only → email. Unambiguous mappings
    # only (an ISRC or title claimed by >1 email is dropped — never guess).
    # Used as fallback attribution tiers when normal artist-name resolution
    # fails but the specific track already exists in Supabase submissions.
    isrc_to_email: dict[str, str] = {}
    _isrc_ambig: dict[str, set[str]] = defaultdict(set)
    for (em, _isrc), _sid in isrc_to_sub.items():
        _isrc_ambig[_isrc].add(em)
    for _isrc, _ems in _isrc_ambig.items():
        if len(_ems) == 1:
            isrc_to_email[_isrc] = next(iter(_ems))

    title_to_email: dict[str, str] = {}
    _title_ambig: dict[str, set[str]] = defaultdict(set)
    for (em, _t), _sid in title_to_sub.items():
        _title_ambig[_t].add(em)
    for _t, _ems in _title_ambig.items():
        if len(_ems) == 1:
            title_to_email[_t] = next(iter(_ems))
    print(f"  isrc_to_email:  {len(isrc_to_email)} unique (Supabase submissions)")
    print(f"  title_to_email: {len(title_to_email)} unique (Supabase submissions)")

    # ---- Aggregate ------------------------------------------------------
    # key: (email, canonical_title, canonical_platform, month_name, year)
    agg: dict[tuple, dict] = {}
    unmatched_artists: dict[str, dict] = {}
    matched_rows = unmatched_rows = skipped_dispute = skipped_no_period = 0
    match_tier = {"isrc": 0, "title": 0, "none": 0}
    # Attribution-tier counters — tracks which tier resolved each row.
    tier_hits: dict[str, int] = defaultdict(int)
    tier_recovered_inr: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    tier_recovered_users: dict[str, set[str]] = defaultdict(set)
    isrcs_in_file: set[str] = set()

    # Pre-compute total xlsx net INR (matched + unmatched) for reconciliation.
    total_xlsx_inr_all = Decimal("0")
    for _r in rows:
        if ((_r.get("adjustment_type") or "").strip() == "Dispute Resolution"):
            continue
        if is_inr_workbook:
            total_xlsx_inr_all += to_decimal(_r.get("royalty_inr") or "0")
        else:
            _period_raw = str(_r.get("source_period") or "").strip()
            if _period_raw in fx_rates:
                total_xlsx_inr_all += to_decimal(_r.get("royalty")) * fx_rates[_period_raw]

    for r in rows:
        # Skip Dispute Resolution adjustment rows entirely.
        if (r.get("adjustment_type") or "").strip() == "Dispute Resolution":
            skipped_dispute += 1
            continue

        period_raw = str(r.get("source_period") or "").strip()
        parsed_period = month_name_from_period(period_raw)
        if parsed_period is None:
            skipped_no_period += 1
            continue
        month_name, year = parsed_period
        if not is_inr_workbook:
            fx = fx_rates[period_raw]

        artist_raw = (r.get("artist") or "").strip()
        sub_label_key = norm(r.get("sub_label"))
        # DSP reports concatenate multiple artists with any of:
        #   * double-space   ("Krantiveer  Zever  Big Bunny")
        #   * comma-space    ("SK Sonu Turi, Laxmi Dubey")
        #   * " feat[uring] " / " ft. " / " x " / " & " / " and "
        # Try (full string) first, then each split candidate, then sub_label.
        # ISRC-match wins as tiebreaker across candidates.
        isrc_raw = (r.get("isrc") or "").strip().upper()
        candidates: list[str] = [norm(artist_raw)]
        # A single regex splits on any of the delimiters. \s+ around " x ",
        # " & ", " and " keeps them from being confused with parts of a name.
        parts = re.split(
            r"\s{2,}|,\s*|\s+(?:featuring|feat\.?|ft\.?|x|&|and)\s+",
            artist_raw,
            flags=re.IGNORECASE,
        )
        for p in parts:
            k = norm(p)
            if k and k not in candidates:
                candidates.append(k)
        if sub_label_key and sub_label_key not in candidates:
            candidates.append(sub_label_key)

        email: str | None = None
        _tier: str | None = None
        # 1. Legacy SQL-dump ISRC map is authoritative (built from
        #    dbo.ReleaseDetails.Isrc → dbo.Users.Email).
        if isrc_raw and isrc_raw in legacy["isrc"]:
            email = legacy["isrc"][isrc_raw]
            _tier = "legacy_isrc"
        # 2. Supabase ISRC-only — an ISRC that unambiguously belongs to a
        #    single Supabase submission owner. Safest tier (ISRCs are
        #    globally unique per-recording).
        if not email and isrc_raw and isrc_raw in isrc_to_email:
            email = isrc_to_email[isrc_raw]
            _tier = "supabase_isrc_only"
        # 3. Prefer a candidate that owns this ISRC (in Supabase submissions).
        if not email and isrc_raw:
            for cand in candidates:
                cand_email = artist_to_email.get(cand)
                if cand_email and (cand_email, isrc_raw) in isrc_to_sub:
                    email = cand_email
                    _tier = "supabase_email_isrc"
                    break
        # 4. First candidate that maps to any known Tunefry user
        #    (Supabase-derived + legacy-derived names).
        if not email:
            for cand in candidates:
                cand_email = artist_to_email.get(cand)
                if cand_email:
                    email = cand_email
                    _tier = "artist_name"
                    break
        # 5. Legacy (artist||title) fallback — catches songs whose artist
        #    display name isn't registered anywhere, but the specific
        #    (artist, title) pair was recorded in dbo.ReleaseDetails.
        if not email:
            title_norm = norm((r.get("track_title") or "").strip())
            for cand in candidates:
                key = f"{cand}||{title_norm}"
                if key in legacy["artist_title"]:
                    email = legacy["artist_title"][key]
                    _tier = "legacy_artist_title"
                    break
        # 6. Supabase title-only — the song title unambiguously belongs to
        #    exactly one Supabase submission owner. Last-resort fallback
        #    before declaring the row unmatched. Ambiguous titles were
        #    dropped from title_to_email at build time.
        if not email:
            title_norm = norm((r.get("track_title") or "").strip())
            if title_norm and title_norm in title_to_email:
                email = title_to_email[title_norm]
                _tier = "supabase_title_only"

        if not email:
            unmatched_rows += 1
            acc = unmatched_artists.setdefault(
                artist_raw,
                {"artist": artist_raw,
                 "sub_label": r.get("sub_label") or "",
                 "row_count": 0,
                 "total_royalty_usd": Decimal("0"),
                 "total_royalty_inr": Decimal("0"),
                 "total_streams": 0,
                 "tracks": {},
                 "platform_data": {},
                 "periods": set()},
            )
            _, _pg = normalize_platform(r.get("source_platform"))
            _row_inr = (to_decimal(r.get("royalty_inr") or "0") if is_inr_workbook
                        else to_decimal(r.get("royalty")) * fx)
            acc["row_count"] += 1
            acc["total_streams"] += to_int(r.get("quantity"))
            acc["total_royalty_inr"] += _row_inr
            acc["total_royalty_usd"] += (Decimal("0") if is_inr_workbook
                                          else to_decimal(r.get("royalty")))
            _pd = acc["platform_data"].setdefault(_pg, {"inr": Decimal("0"), "streams": 0})
            _pd["inr"] += _row_inr
            _pd["streams"] += to_int(r.get("quantity"))
            acc["periods"].add(period_raw)
            _tk_key = ((r.get("track_title") or "").strip(), isrc_raw)
            _tk = acc["tracks"].setdefault(_tk_key, {
                "title": (r.get("track_title") or "").strip(),
                "isrc": isrc_raw, "streams": 0, "inr": Decimal("0"),
            })
            _tk["streams"] += to_int(r.get("quantity"))
            _tk["inr"] += _row_inr
            continue
        matched_rows += 1

        title_raw = (r.get("track_title") or "").strip()
        canonical_title = title_raw  # display-as-is (no case fold)
        canonical_platform, group = normalize_platform(r.get("source_platform"))
        streams = to_int(r.get("quantity"))
        if is_inr_workbook:
            revenue_inr = to_decimal(r.get("royalty_inr") or "0")
        else:
            royalty_usd = to_decimal(r.get("royalty"))
            revenue_inr = royalty_usd * fx

        if _tier:
            tier_hits[_tier] += 1
            tier_recovered_inr[_tier] += revenue_inr
            tier_recovered_users[_tier].add(email)

        # submission_id linkage — ISRC preferred, then (title) fallback.
        isrc = (r.get("isrc") or "").strip().upper()
        sub_id: str | None = None
        if isrc:
            isrcs_in_file.add(isrc)
            sub_id = isrc_to_sub.get((email, isrc))
            if sub_id:
                match_tier["isrc"] += 1
        if not sub_id:
            sub_id = title_to_sub.get((email, norm(title_raw)))
            if sub_id:
                match_tier["title"] += 1
        if not sub_id:
            match_tier["none"] += 1

        key = (email, canonical_title, canonical_platform, month_name, year)
        acc = agg.get(key)
        if acc is None:
            acc = {
                "user_email": email,
                "submission_id": sub_id,
                "artist_name": artist_raw,
                "song_title": canonical_title,
                "platform": canonical_platform,
                "platform_group": group,
                "period_month": month_name,
                "period_year": year,
                "streams": 0,
                "revenue": Decimal("0"),
            }
            agg[key] = acc
        if sub_id and not acc["submission_id"]:
            acc["submission_id"] = sub_id
        acc["streams"] += streams
        acc["revenue"] += revenue_inr

    print(f"  matched_rows={matched_rows}  unmatched_rows={unmatched_rows}  "
          f"skipped(dispute)={skipped_dispute}  skipped(bad_period)={skipped_no_period}")
    print(f"  -> {len(agg)} song_stats rows after aggregation")
    print(f"  submission_id match tiers: "
          f"isrc={match_tier['isrc']}  title={match_tier['title']}  "
          f"none={match_tier['none']}")

    # ---- Aggregate CSV rows (Apple Music, already INR — no FX) ----------
    csv_agg: dict[tuple, dict] = {}
    if csv_rows and csv_period:
        # Keep a stable reference to the legacy map; the balance loop later
        # reassigns the name `legacy` to a Decimal per-user.
        _legacy_map = legacy
        csv_parsed = month_name_from_period(csv_period)
        csv_month_name, csv_year_int = csv_parsed  # type: ignore[misc]
        csv_matched_count = csv_unmatched_count = 0
        csv_match_tier: dict[str, int] = {"isrc": 0, "title": 0, "none": 0}

        for r in csv_rows:
            artist_raw = (r.get("artist") or "").strip()
            sub_label_key = norm(r.get("sub_label"))
            isrc_raw = r["isrc"]
            candidates: list[str] = [norm(artist_raw)]
            parts = re.split(
                r"\s{2,}|,\s*|\s+(?:featuring|feat\.?|ft\.?|x|&|and)\s+",
                artist_raw, flags=re.IGNORECASE,
            )
            for p in parts:
                k = norm(p)
                if k and k not in candidates:
                    candidates.append(k)
            if sub_label_key and sub_label_key not in candidates:
                candidates.append(sub_label_key)

            email: str | None = None
            _tier: str | None = None
            if isrc_raw and isrc_raw in _legacy_map["isrc"]:
                email = _legacy_map["isrc"][isrc_raw]
                _tier = "legacy_isrc"
            if not email and isrc_raw and isrc_raw in isrc_to_email:
                email = isrc_to_email[isrc_raw]
                _tier = "supabase_isrc_only"
            if not email and isrc_raw:
                for cand in candidates:
                    cand_email = artist_to_email.get(cand)
                    if cand_email and (cand_email, isrc_raw) in isrc_to_sub:
                        email = cand_email
                        _tier = "supabase_email_isrc"
                        break
            if not email:
                for cand in candidates:
                    cand_email = artist_to_email.get(cand)
                    if cand_email:
                        email = cand_email
                        _tier = "artist_name"
                        break
            if not email:
                _title_norm = norm((r.get("track_title") or "").strip())
                for cand in candidates:
                    _key_at = f"{cand}||{_title_norm}"
                    if _key_at in _legacy_map["artist_title"]:
                        email = _legacy_map["artist_title"][_key_at]
                        _tier = "legacy_artist_title"
                        break
            if not email:
                _title_norm = norm((r.get("track_title") or "").strip())
                if _title_norm and _title_norm in title_to_email:
                    email = title_to_email[_title_norm]
                    _tier = "supabase_title_only"

            if not email:
                csv_unmatched_count += 1
                acc = unmatched_artists.setdefault(
                    artist_raw,
                    {"artist": artist_raw,
                     "sub_label": r.get("sub_label") or "",
                     "row_count": 0,
                     "total_royalty_usd": Decimal("0"),
                     "total_royalty_inr": Decimal("0"),
                     "total_streams": 0,
                     "tracks": {},
                     "platform_data": {},
                     "periods": set()},
                )
                _row_inr = r["_royalty_inr"]
                acc["row_count"] += 1
                acc["total_streams"] += to_int(r.get("quantity"))
                acc["total_royalty_inr"] += _row_inr
                _pd = acc["platform_data"].setdefault("Apple Music", {"inr": Decimal("0"), "streams": 0})
                _pd["inr"] += _row_inr
                _pd["streams"] += to_int(r.get("quantity"))
                acc["periods"].add(csv_period)
                _tk_key = ((r.get("track_title") or "").strip(), r["isrc"])
                _tk = acc["tracks"].setdefault(_tk_key, {
                    "title": (r.get("track_title") or "").strip(),
                    "isrc": r["isrc"], "streams": 0, "inr": Decimal("0"),
                })
                _tk["streams"] += to_int(r.get("quantity"))
                _tk["inr"] += _row_inr
                continue
            csv_matched_count += 1

            title_raw = (r.get("track_title") or "").strip()
            streams = to_int(r.get("quantity"))
            revenue_inr = r["_royalty_inr"]  # already INR — NO FX conversion

            if _tier:
                tier_hits[_tier] += 1
                tier_recovered_inr[_tier] += revenue_inr
                tier_recovered_users[_tier].add(email)

            sub_id: str | None = None
            if isrc_raw:
                sub_id = isrc_to_sub.get((email, isrc_raw))
                if sub_id:
                    csv_match_tier["isrc"] += 1
            if not sub_id:
                sub_id = title_to_sub.get((email, norm(title_raw)))
                if sub_id:
                    csv_match_tier["title"] += 1
            if not sub_id:
                csv_match_tier["none"] += 1

            key = (email, title_raw, "Apple Music", csv_month_name, csv_year_int)
            acc = csv_agg.get(key)
            if acc is None:
                acc = {
                    "user_email": email,
                    "submission_id": sub_id,
                    "artist_name": artist_raw,
                    "song_title": title_raw,
                    "platform": "Apple Music",
                    "platform_group": "Apple Music",
                    "period_month": csv_month_name,
                    "period_year": csv_year_int,
                    "streams": 0,
                    "revenue": Decimal("0"),
                }
                csv_agg[key] = acc
            if sub_id and not acc["submission_id"]:
                acc["submission_id"] = sub_id
            acc["streams"] += streams
            acc["revenue"] += revenue_inr

        print(f"\n  Apple Music CSV: matched={csv_matched_count}  "
              f"unmatched={csv_unmatched_count}  -> {len(csv_agg)} rows")
        print(f"  CSV sub_id tiers: isrc={csv_match_tier['isrc']}  "
              f"title={csv_match_tier['title']}  none={csv_match_tier['none']}")

        # Merge: CSV fully replaces xlsx Apple Music rows for this period to
        # prevent double-counting (xlsx and CSV both report Apple Music).
        before_merge = len(agg)
        agg = {k: v for k, v in agg.items()
               if not (k[2] == "Apple Music"
                       and k[3] == csv_month_name
                       and k[4] == csv_year_int)}
        dropped_xlsx_am = before_merge - len(agg)
        agg.update(csv_agg)
        print(f"  Merge: dropped {dropped_xlsx_am} xlsx Apple Music {csv_month_name} "
              f"{csv_year_int} rows; added {len(csv_agg)} from CSV. "
              f"Final agg: {len(agg)} rows.")

    # ---- Attribution-tier breakdown -------------------------------------
    print(f"\n-- Attribution tiers (rows / INR / distinct users) --")
    _tier_order = [
        "legacy_isrc",
        "supabase_isrc_only",
        "supabase_email_isrc",
        "artist_name",
        "legacy_artist_title",
        "supabase_title_only",
    ]
    for _name in _tier_order:
        _rows = tier_hits.get(_name, 0)
        _inr = tier_recovered_inr.get(_name, Decimal("0"))
        _users = len(tier_recovered_users.get(_name, set()))
        marker = "  [NEW]" if _name in ("supabase_isrc_only", "supabase_title_only") else ""
        print(f"  {_name:<24s}: rows={_rows:<6d}  ₹{_inr:>12.2f}  users={_users}{marker}")

    # ---- Reconciliation check -------------------------------------------
    def _parse_inr(s: str | None) -> Decimal | None:
        if not s:
            return None
        cleaned = re.sub(r"[₹,\s]", "", s)
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            print(f"  WARNING: could not parse expected amount {s!r} — skipping check.")
            return None

    total_to_write_inr = sum(acc["revenue"] for acc in agg.values())
    total_xlsx_unmatched_inr = sum(
        (a.get("total_royalty_inr") or Decimal("0"))
        for a in unmatched_artists.values()
    )
    exp_net = _parse_inr(args.expected_net)
    exp_gross = _parse_inr(args.expected_gross)

    print(f"\n-- Reconciliation --")
    print(f"  xlsx net total (all rows after FX):    ₹{total_xlsx_inr_all:.2f}")
    if csv_rows:
        print(f"  Apple Music CSV net total (INR):       ₹{total_csv_inr_all:.2f}")
    print(f"  Unattributed xlsx revenue:             ₹{total_xlsx_unmatched_inr:.2f}")
    print(f"  Total to write to song_stats:          ₹{total_to_write_inr:.2f}")
    if exp_gross is not None:
        diff = total_xlsx_inr_all - exp_gross
        status = "OK" if abs(diff) <= Decimal("1") else f"WARNING diff ₹{diff:.2f}"
        print(f"  Expected gross/net (xlsx):             ₹{exp_gross:.2f}  [{status}]")
    if exp_net is not None:
        if total_to_write_inr > exp_net + Decimal("1"):
            msg = (f"\n  AMOUNT CAP BREACH: about to write ₹{total_to_write_inr:.2f} "
                   f"but expected net is ₹{exp_net:.2f}. "
                   f"Over by ₹{(total_to_write_inr - exp_net):.2f}. "
                   f"Check FX rates / attribution.")
            print(msg, file=sys.stderr)
            if not dry_run:
                print("  ABORTING live write to protect real balances.", file=sys.stderr)
                sys.exit(1)
            else:
                print("  (dry-run — no writes occurred; fix before --live)")
        else:
            pct = (total_to_write_inr / exp_net * 100) if exp_net > 0 else Decimal("0")
            print(f"  Expected net cap:                      ₹{exp_net:.2f}  "
                  f"writing ₹{total_to_write_inr:.2f} ({pct:.1f}% of cap)  [OK]")

    # ---- Write unmatched CSV (always — dry-run or not) ------------------
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M")
    if unmatched_artists:
        csv_path = xlsx_path.parent / f"unmatched_{ts}.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["artist", "sub_label", "row_count",
                        "total_royalty_usd", "total_royalty_inr", "total_streams"])
            sort_key = (
                (lambda x: -(float(x.get("total_royalty_inr") or 0)))
                if is_inr_workbook
                else (lambda x: -(float(x.get("total_royalty_usd") or 0)))
            )
            for a in sorted(unmatched_artists.values(), key=sort_key):
                w.writerow([a["artist"], a["sub_label"], a["row_count"],
                            str(a["total_royalty_usd"]),
                            str(a.get("total_royalty_inr") or "0"),
                            a["total_streams"]])
        print(f"  Unmatched artists -> {csv_path}  "
              f"({len(unmatched_artists)} distinct)")
        xl_path = csv_path.with_suffix(".xlsx")
        _write_unmatched_excel(xl_path, unmatched_artists, is_inr_workbook)
        print(f"  Unmatched Excel  -> {xl_path}")

    # ---- Emails covered by this file ------------------------------------
    emails_in_file: set[str] = {acc["user_email"] for acc in agg.values()}
    print(f"  {len(emails_in_file)} distinct users in this file")

    # ---- Compute new artist_balances ------------------------------------
    # Existing balance users get recomputed too, so a paid request approved
    # since last ingest lowers their available_balance correctly.
    existing_emails = list_existing_balance_emails(svc)
    all_emails = emails_in_file | existing_emails

    print("Loading current paid/pending withdrawal_requests…")
    paid, pending = load_reserved_withdrawals(svc)
    print(f"  paid users={len(paid)}  pending users={len(pending)}")

    # For dry-run, we compute total_earned by summing:
    #   existing song_stats (for periods NOT covered by this file)  +
    #   NEW aggregated rows (for periods covered by this file)
    # For live-run we do the same in-memory calc, then delete+insert covered
    # periods and upsert balances in one shot.
    print("Loading existing song_stats (all rows for affected users)…")
    existing_by_email: dict[str, list[dict]] = defaultdict(list)
    if all_emails:
        # Fetch in batches of 200 emails via IN filter.
        emails_list = list(all_emails)
        for i in range(0, len(emails_list), 200):
            chunk = emails_list[i:i + 200]
            offset = 0
            while True:
                res = (svc.table("song_stats")
                       .select("user_email, period_month, period_year, revenue, streams")
                       .in_("user_email", chunk)
                       .range(offset, offset + 999).execute())
                page = res.data or []
                for r in page:
                    existing_by_email[(r.get("user_email") or "").lower()].append(r)
                if len(page) < 1000:
                    break
                offset += 1000
    print(f"  loaded existing song_stats for {len(existing_by_email)} users")

    # Build per-user "before" totals (from existing rows, excluding covered periods).
    covered_periods: set[tuple[str, int]] = {
        (acc["period_month"], acc["period_year"]) for acc in agg.values()
    }
    print(f"  covered periods: {sorted(covered_periods)}")

    total_earned_before: dict[str, Decimal] = defaultdict(Decimal)
    for email, srows in existing_by_email.items():
        for sr in srows:
            key = ((sr.get("period_month") or ""), int(sr.get("period_year") or 0))
            if key in covered_periods:
                # This will be replaced — don't count in "before" total.
                continue
            total_earned_before[email] += to_decimal(sr.get("revenue"))

    # Add new revenue per user.
    new_revenue_by_email: dict[str, Decimal] = defaultdict(Decimal)
    for acc in agg.values():
        new_revenue_by_email[acc["user_email"]] += acc["revenue"]

    # Final total_earned = kept_from_existing + new
    total_earned_after: dict[str, Decimal] = defaultdict(Decimal)
    for email in all_emails:
        total_earned_after[email] = (
            total_earned_before.get(email, Decimal("0"))
            + new_revenue_by_email.get(email, Decimal("0"))
        )

    # Balance computation
    balances: list[dict] = []
    per_user_deltas: list[tuple[str, Decimal, Decimal]] = []  # (email, before, after)
    for email in sorted(all_emails):
        earned = total_earned_after[email]
        legacy = baseline.get(email, Decimal("0"))
        paid_now = paid.get(email, Decimal("0"))
        pending_now = pending.get(email, Decimal("0"))
        withdrawn = legacy + paid_now
        avail = earned - withdrawn - pending_now
        if avail < 0:
            avail = Decimal("0")
        balances.append({
            "user_email": email,
            "total_earned": str(earned),
            "total_withdrawn": str(withdrawn),
            "available_balance": str(avail),
        })
        prev_earned = sum(
            (to_decimal(sr.get("revenue"))
             for sr in existing_by_email.get(email, [])),
            Decimal("0"),
        )
        per_user_deltas.append((email, prev_earned, earned))

    # Submissions to approve (unique sub_ids the file linked to).
    approve_ids: set[str] = {acc["submission_id"] for acc in agg.values()
                             if acc["submission_id"]}

    # ---- Dry-run report -------------------------------------------------
    tag = "[DRY RUN] " if dry_run else ""
    print(f"\n{tag}Plan:")
    print(f"  * delete + insert song_stats for {len(emails_in_file)} users "
          f"across {len(covered_periods)} periods -> {len(agg)} new rows")
    print(f"  * upsert artist_balances for {len(balances)} users")
    print(f"  * approve up to {len(approve_ids)} submissions "
          f"(only if currently != approved)")

    # Top 10 balance movers (biggest earned delta).
    movers = sorted(per_user_deltas, key=lambda t: (t[2] - t[1]), reverse=True)[:10]
    print("\nTop 10 total_earned deltas:")
    print(f"  {'email':<40} {'before':>14} {'after':>14} {'delta':>14}")
    for email, before, after in movers:
        delta = after - before
        print(f"  {email:<40} {str(before):>14} {str(after):>14} {str(delta):>14}")

    if dry_run:
        print("\nDRY RUN -- nothing written. Pass --live to write.")
        return

    # ---- LIVE WRITES ----------------------------------------------------
    # Order:
    #   1. Delete existing song_stats for covered periods per matched user
    #   2. Insert new song_stats rows (chunked)
    #   3. Upsert artist_balances
    #   4. Flip submission status to approved (chunked)
    print("\n== LIVE WRITE PHASE ==")
    now = datetime.now(timezone.utc).isoformat()

    # 1. Period-replace delete.
    # Per matched user, delete existing rows for each covered (month, year).
    # We use exact eq() per period (rather than IN over months) so we can
    # never accidentally touch a period this file doesn't cover.
    emails_list = sorted(emails_in_file)
    total_deleted = 0
    for i in range(0, len(emails_list), DELETE_CHUNK):
        chunk = emails_list[i:i + DELETE_CHUNK]
        for month_name, year in covered_periods:
            res = (svc.table("song_stats").delete()
                   .in_("user_email", chunk)
                   .eq("period_month", month_name)
                   .eq("period_year", year)
                   .execute())
            total_deleted += len(res.data or [])
    print(f"  deleted {total_deleted} existing song_stats rows "
          f"(periods={sorted(covered_periods)})")

    # Extra delete for CSV period when it isn't covered by the xlsx period-replace.
    # Example: xlsx covers Feb + Apr only, but csv_period is March — the xlsx delete
    # didn't touch March, so any stale Apple Music March rows must be removed now.
    if csv_agg and csv_period:
        _csv_month_del, _csv_year_del = month_name_from_period(csv_period)  # type: ignore[misc]
        if (_csv_month_del, _csv_year_del) not in covered_periods:
            csv_emails_del = sorted({acc["user_email"] for acc in csv_agg.values()})
            for i in range(0, len(csv_emails_del), DELETE_CHUNK):
                chunk = csv_emails_del[i:i + DELETE_CHUNK]
                res = (svc.table("song_stats").delete()
                       .in_("user_email", chunk)
                       .eq("platform", "Apple Music")
                       .eq("period_month", _csv_month_del)
                       .eq("period_year", _csv_year_del)
                       .execute())
                total_deleted += len(res.data or [])
            print(f"  + deleted stale Apple Music {_csv_month_del} {_csv_year_del} rows "
                  f"(CSV-only period, platform-scoped delete)")

    # 2. Insert new rows.
    insert_rows = []
    for acc in agg.values():
        insert_rows.append({
            "user_email": acc["user_email"],
            "submission_id": acc["submission_id"],
            "artist_name": acc["artist_name"],
            "song_title": acc["song_title"],
            "platform": acc["platform"],
            "platform_group": acc["platform_group"],
            "period_month": acc["period_month"],
            "period_year": acc["period_year"],
            "streams": acc["streams"],
            "revenue": str(acc["revenue"]),
            "updated_at": now,
        })
    inserted = 0
    for i in range(0, len(insert_rows), UPSERT_BATCH):
        chunk = insert_rows[i:i + UPSERT_BATCH]
        # Use upsert on the UNIQUE key so aggregation collisions inside this
        # file (should never happen, but defensive) still succeed.
        svc.table("song_stats").upsert(
            chunk,
            on_conflict="user_email,song_title,platform,period_month,period_year",
        ).execute()
        inserted += len(chunk)
        print(f"  inserted song_stats {inserted}/{len(insert_rows)}")

    # 3. Upsert artist_balances.
    for b in balances:
        b["last_updated"] = now
    for i in range(0, len(balances), UPSERT_BATCH):
        chunk = balances[i:i + UPSERT_BATCH]
        svc.table("artist_balances").upsert(chunk, on_conflict="user_email").execute()
    print(f"  upserted {len(balances)} artist_balances")

    # 4. Flip matched submissions to approved.
    ids = list(approve_ids)
    flipped = 0
    for i in range(0, len(ids), DELETE_CHUNK):
        chunk = ids[i:i + DELETE_CHUNK]
        res = (svc.table("submissions")
               .update({"status": "approved"})
               .in_("id", chunk).neq("status", "approved").execute())
        flipped += len(res.data or [])
    print(f"  marked {flipped} submissions approved")

    print("\nDone.")


if __name__ == "__main__":
    main()
